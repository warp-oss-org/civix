"""Tests for the England FCERM source slice."""

from __future__ import annotations

from collections.abc import Callable
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
import pytest
import respx
from openpyxl import Workbook, load_workbook

from civix.core.drift import SchemaObserver, TaxonomyDriftKind, TaxonomyObserver
from civix.core.identity.models.identifiers import SnapshotId
from civix.core.pipeline import attach_observers, run
from civix.core.ports.errors import FetchError
from civix.core.ports.models.adapter import SourceAdapter
from civix.core.quality.models.fields import FieldQuality
from civix.core.snapshots.models.snapshot import RawRecord, SourceSnapshot
from civix.core.temporal import TemporalPeriodPrecision
from civix.domains.hazard_mitigation.adapters.sources.gb.fcerm import (
    DEFAULT_RESOURCE_URL,
    ENGLAND_FCERM_PROGRAMME_YEAR,
    ENGLAND_FCERM_PUBLICATION_URL,
    ENGLAND_FCERM_SCHEMES_DATASET_ID,
    ENGLAND_FCERM_SCHEMES_SCHEMA,
    ENGLAND_FCERM_SCHEMES_SHEET_NAME,
    ENGLAND_FCERM_SCHEMES_TAXONOMIES,
    ENGLAND_FCERM_SOURCE_RECORD_ID_POLICY,
    ENGLAND_FCERM_SOURCE_SCOPE,
    GB_ENGLAND_JURISDICTION,
    INDICATIVE_GOVERNMENT_INVESTMENT_FIELD,
    PROJECT_TYPE_FIELD,
    RFCC_FIELD,
    RISK_SOURCE_FIELD,
    SOURCE_ID,
    EnglandFcermCaveat,
    EnglandFcermFetchConfig,
    EnglandFcermProjectMapper,
    EnglandFcermSchemesAdapter,
)
from civix.domains.hazard_mitigation.models.common import (
    MitigationFundingAmountKind,
    MitigationFundingEventType,
    MitigationFundingShareKind,
    MitigationHazardType,
    MitigationInterventionType,
)

PINNED_NOW = datetime(2026, 5, 3, 12, 0, tzinfo=UTC)
FIXTURES = Path(__file__).parent / "fixtures"
XLSX_FIXTURE = FIXTURES / "schemes_2026_2027.xlsx"


def _xlsx_bytes(path: Path = XLSX_FIXTURE) -> bytes:
    return path.read_bytes()


def _fetch(client: httpx.AsyncClient) -> EnglandFcermFetchConfig:
    return EnglandFcermFetchConfig(client=client, clock=lambda: PINNED_NOW)


def _adapter(client: httpx.AsyncClient) -> EnglandFcermSchemesAdapter:
    return EnglandFcermSchemesAdapter(fetch_config=_fetch(client))


def _snapshot() -> SourceSnapshot:
    return SourceSnapshot(
        snapshot_id=SnapshotId("england-fcerm-snapshot"),
        source_id=SOURCE_ID,
        dataset_id=ENGLAND_FCERM_SCHEMES_DATASET_ID,
        jurisdiction=GB_ENGLAND_JURISDICTION,
        fetched_at=PINNED_NOW,
        record_count=1,
    )


def _record(raw: dict[str, Any], source_record_id: str) -> RawRecord:
    return RawRecord(
        snapshot_id=SnapshotId("england-fcerm-snapshot"),
        raw_data=raw,
        source_record_id=source_record_id,
    )


def _rows() -> list[dict[str, Any]]:
    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    worksheet = workbook[ENGLAND_FCERM_SCHEMES_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    decoded = [dict(zip(headers, row, strict=True)) for row in rows[1:]]
    workbook.close()

    return decoded


def _capture_sequence(
    requests: list[httpx.Request],
    responses: list[httpx.Response],
) -> Callable[[httpx.Request], httpx.Response]:
    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)

        return responses.pop(0)

    return handler


def _remove_sheet_payload(sheet_name: str) -> bytes:
    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    del workbook[sheet_name]
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def _bad_header_payload() -> bytes:
    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    worksheet = workbook[ENGLAND_FCERM_SCHEMES_SHEET_NAME]
    worksheet.cell(row=1, column=1, value="Scheme Name")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def _bad_row_payload() -> bytes:
    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    worksheet = workbook[ENGLAND_FCERM_SCHEMES_SHEET_NAME]
    worksheet.cell(row=2, column=1, value="")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def _blank_row_between_schemes_payload() -> bytes:
    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    worksheet = workbook[ENGLAND_FCERM_SCHEMES_SHEET_NAME]
    worksheet.insert_rows(3)
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def _blank_workbook_payload() -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.create_sheet(ENGLAND_FCERM_SCHEMES_SHEET_NAME)

    if default_sheet is not None:
        workbook.remove(default_sheet)

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


class TestAdapter:
    async def test_fetches_xlsx_and_preserves_source_shape(self) -> None:
        requests: list[httpx.Request] = []

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(DEFAULT_RESOURCE_URL).mock(
                side_effect=_capture_sequence(
                    requests,
                    [httpx.Response(200, content=_xlsx_bytes())],
                )
            )

            async with httpx.AsyncClient() as client:
                adapter = _adapter(client)
                result = await adapter.fetch()
                records = [record async for record in result.records]

        assert isinstance(adapter, SourceAdapter)
        assert requests[0].url == DEFAULT_RESOURCE_URL
        assert result.snapshot.source_id == SOURCE_ID
        assert result.snapshot.dataset_id == ENGLAND_FCERM_SCHEMES_DATASET_ID
        assert result.snapshot.jurisdiction == GB_ENGLAND_JURISDICTION
        assert result.snapshot.record_count == 4
        assert result.snapshot.source_url == DEFAULT_RESOURCE_URL
        assert result.snapshot.content_hash is not None
        assert result.snapshot.fetch_params == {
            "publication_url": ENGLAND_FCERM_PUBLICATION_URL,
            "programme_year": ENGLAND_FCERM_PROGRAMME_YEAR,
            "resource_url": DEFAULT_RESOURCE_URL,
            "sheet_name": ENGLAND_FCERM_SCHEMES_SHEET_NAME,
            "source_record_id_policy": ENGLAND_FCERM_SOURCE_RECORD_ID_POLICY,
        }

        assert [record.source_record_id for record in records] == [
            "abingdon-partnership-flood-alleviation-scheme:environment-agency:oxfordshire:row-2",
            (
                "adaptive-scillies-natural-dune-restoration-and-flood-resilience:"
                "council-of-the-isles-of-scilly:cornwall:row-3"
            ),
            ("alfold-crossways-flood-alleviation-scheme:surrey-county-council:surrey:row-4"),
            "bispham-coast-protection-scheme:environment-agency:lancashire:row-5",
        ]

        assert records[0].raw_data[INDICATIVE_GOVERNMENT_INVESTMENT_FIELD] == 30
        assert records[0].record_hash is not None

    async def test_source_record_ids_preserve_original_workbook_row_numbers(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(DEFAULT_RESOURCE_URL).mock(
                return_value=httpx.Response(200, content=_blank_row_between_schemes_payload())
            )

            async with httpx.AsyncClient() as client:
                result = await _adapter(client).fetch()
                records = [record async for record in result.records]

        assert result.snapshot.record_count == 4
        assert [record.source_record_id for record in records] == [
            "abingdon-partnership-flood-alleviation-scheme:environment-agency:oxfordshire:row-2",
            (
                "adaptive-scillies-natural-dune-restoration-and-flood-resilience:"
                "council-of-the-isles-of-scilly:cornwall:row-4"
            ),
            ("alfold-crossways-flood-alleviation-scheme:surrey-county-council:surrey:row-5"),
            "bispham-coast-protection-scheme:environment-agency:lancashire:row-6",
        ]

    @pytest.mark.parametrize(
        ("payload_factory", "message"),
        [
            (lambda: b"not an xlsx", "invalid FCERM XLSX"),
            (lambda: _remove_sheet_payload(ENGLAND_FCERM_SCHEMES_SHEET_NAME), "missing"),
            (_bad_header_payload, "header"),
            (_blank_workbook_payload, "header"),
            (_bad_row_payload, "missing required field"),
        ],
    )
    async def test_rejects_malformed_xlsx_payloads(
        self,
        payload_factory: Callable[[], bytes],
        message: str,
    ) -> None:
        payload = payload_factory()

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(DEFAULT_RESOURCE_URL).mock(
                return_value=httpx.Response(200, content=payload)
            )

            async with httpx.AsyncClient() as client:
                with pytest.raises(FetchError, match=message):
                    await _adapter(client).fetch()


class TestMapper:
    def test_maps_complete_scheme_project(self) -> None:
        result = EnglandFcermProjectMapper()(_record(_rows()[0], "fcerm-row-1"), _snapshot())
        project = result.record

        assert project.project_id == "fcerm-row-1"
        assert project.title.value == "Abingdon Partnership Flood Alleviation Scheme"
        assert project.description.quality is FieldQuality.UNMAPPED
        assert project.programme.value is not None
        assert project.programme.value.code == "fcerm-2026-27"
        assert project.organizations.value is not None
        assert project.organizations.value[0].name == "Environment Agency"
        assert project.hazard_types.value == (MitigationHazardType.FLOOD,)
        assert project.source_hazards.value is not None
        assert project.source_hazards.value[0].label == "River Flooding"
        assert project.intervention_types.value == (MitigationInterventionType.FLOOD_DEFENCE,)
        assert project.source_interventions.value is not None
        assert project.source_interventions.value[0].label == "Defence"
        assert project.status.quality is FieldQuality.UNMAPPED

    def test_maps_geography_fiscal_period_funding_and_caveats(self) -> None:
        project = EnglandFcermProjectMapper()(
            _record(_rows()[0], "fcerm-row-1"),
            _snapshot(),
        ).record

        assert project.fiscal_period.value is not None
        assert project.fiscal_period.value.precision is TemporalPeriodPrecision.INTERVAL
        assert project.fiscal_period.value.start_datetime == datetime(2026, 4, 1)
        assert project.fiscal_period.value.end_datetime == datetime(2027, 3, 31)
        assert project.geography.value is not None
        assert project.geography.value[0].place_name == "Oxfordshire"
        assert project.geography.value[0].footprint is None
        assert project.geography.value[0].administrative_areas == (
            "South East",
            "Thames",
            "Oxford West and Abingdon",
            "Oxfordshire",
        )

        assert project.funding_summaries.value is not None
        assert project.funding_summaries.value[0].money.amount == Decimal("30000")
        assert project.funding_summaries.value[0].money.currency == "GBP"
        assert project.funding_summaries.value[0].amount_kind is (
            MitigationFundingAmountKind.PROJECT_AMOUNT
        )

        assert project.funding_summaries.value[0].share_kind is (
            MitigationFundingShareKind.GOVERNMENT
        )

        assert project.funding_summaries.value[0].lifecycle is (
            MitigationFundingEventType.PLANNED_AMOUNT
        )

        assert project.source_caveats.value is not None
        assert {caveat.code for caveat in project.source_caveats.value} == {
            "compiled-information",
            "funding-and-timelines-change",
            "future-funding-review",
            "live-projects",
            "row-derived-identifiers",
            "scheme-level-location",
            "year-only-investment",
        }

    def test_property_flood_resilience_remains_source_specific(self) -> None:
        project = EnglandFcermProjectMapper()(
            _record(_rows()[2], "fcerm-row-3"),
            _snapshot(),
        ).record

        assert project.hazard_types.value == (MitigationHazardType.FLOOD,)
        assert project.intervention_types.value == (MitigationInterventionType.SOURCE_SPECIFIC,)

    def test_coastal_erosion_maps_explicitly(self) -> None:
        project = EnglandFcermProjectMapper()(
            _record(_rows()[3], "fcerm-row-4"),
            _snapshot(),
        ).record

        assert project.hazard_types.value == (MitigationHazardType.COASTAL_EROSION,)

    def test_unmapped_source_field_report_is_clean_for_fixture_fields(self) -> None:
        report = EnglandFcermProjectMapper()(_record(_rows()[0], "fcerm-row-1"), _snapshot()).report

        assert report.unmapped_source_fields == ()


class TestDrift:
    async def test_fcerm_fixture_drift_clean(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(DEFAULT_RESOURCE_URL).mock(
                return_value=httpx.Response(200, content=_xlsx_bytes())
            )

            async with httpx.AsyncClient() as client:
                pipeline_result = await run(_adapter(client), EnglandFcermProjectMapper())
                schema_obs = SchemaObserver(spec=ENGLAND_FCERM_SCHEMES_SCHEMA)
                taxonomy_obs = TaxonomyObserver(specs=ENGLAND_FCERM_SCHEMES_TAXONOMIES)
                observed = attach_observers(pipeline_result, [schema_obs, taxonomy_obs])
                records = [record async for record in observed.records]

        assert len(records) == 4
        assert schema_obs.finalize(pipeline_result.snapshot).findings == ()
        assert taxonomy_obs.finalize(pipeline_result.snapshot).findings == ()

    async def test_unknown_taxonomies_surface_as_drift(self) -> None:
        workbook = load_workbook(XLSX_FIXTURE, data_only=True)
        worksheet = workbook[ENGLAND_FCERM_SCHEMES_SHEET_NAME]
        headers = [str(value) for value in next(worksheet.iter_rows(max_row=1, values_only=True))]
        worksheet.cell(
            row=2,
            column=headers.index(RISK_SOURCE_FIELD) + 1,
            value="Reservoir Flooding",
        )
        worksheet.cell(row=2, column=headers.index(PROJECT_TYPE_FIELD) + 1, value="New Type")
        worksheet.cell(row=2, column=headers.index(RFCC_FIELD) + 1, value="New RFCC")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(DEFAULT_RESOURCE_URL).mock(
                return_value=httpx.Response(200, content=output.getvalue())
            )

            async with httpx.AsyncClient() as client:
                pipeline_result = await run(_adapter(client), EnglandFcermProjectMapper())
                taxonomy_obs = TaxonomyObserver(specs=ENGLAND_FCERM_SCHEMES_TAXONOMIES)
                observed = attach_observers(pipeline_result, [taxonomy_obs])
                async for _ in observed.records:
                    pass

        report = taxonomy_obs.finalize(pipeline_result.snapshot)

        assert any(
            finding.kind is TaxonomyDriftKind.UNRECOGNIZED_VALUE
            and finding.taxonomy_id == "england-fcerm-risk-source"
            for finding in report.findings
        )

        assert any(
            finding.kind is TaxonomyDriftKind.UNRECOGNIZED_VALUE
            and finding.taxonomy_id == "england-fcerm-project-type"
            for finding in report.findings
        )

        assert any(
            finding.kind is TaxonomyDriftKind.UNRECOGNIZED_VALUE
            and finding.taxonomy_id == "england-fcerm-rfcc"
            for finding in report.findings
        )


def test_source_metadata_preserves_scope_and_ids() -> None:
    assert SOURCE_ID == "environment-agency"
    assert ENGLAND_FCERM_SCHEMES_DATASET_ID == "fcerm-schemes-2026-2027"
    assert "rotated asset URL" in ENGLAND_FCERM_SOURCE_SCOPE
    assert "row reordering can churn ids" in ENGLAND_FCERM_SOURCE_RECORD_ID_POLICY
    assert EnglandFcermCaveat.ROW_DERIVED_IDENTIFIERS.value
