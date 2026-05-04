"""Tests for the Ontario EWRB source slice."""

from __future__ import annotations

from collections.abc import Callable
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
import pytest
import respx
from openpyxl import Workbook, load_workbook

from civix.core.drift import (
    SchemaDriftKind,
    SchemaObserver,
    TaxonomyDriftKind,
    TaxonomyObserver,
)
from civix.core.identity.models.identifiers import SnapshotId
from civix.core.mapping.errors import MappingError
from civix.core.pipeline import attach_observers, run
from civix.core.ports.errors import FetchError
from civix.core.ports.models.adapter import SourceAdapter
from civix.core.quality.models.fields import FieldQuality
from civix.core.snapshots.models.snapshot import RawRecord, SourceSnapshot
from civix.core.temporal import TemporalPeriodPrecision
from civix.domains.building_energy_emissions.adapters.sources.ca.ontario_ewrb import (
    EWRB_CAVEAT_TAXONOMY_ID,
    EWRB_DATASET_ID,
    EWRB_DEFAULT_REPORTING_YEAR,
    EWRB_DEFAULT_SHEET_NAME,
    EWRB_DEFAULT_URL,
    EWRB_RAW_SCHEMA,
    EWRB_REPORTING_YEAR_FIELD,
    EWRB_SOURCE_SCOPE,
    EWRB_TAXONOMIES,
    METHODOLOGY_VERSION_POST_AUG_2023,
    METHODOLOGY_VERSION_PRE_AUG_2023,
    ONTARIO_JURISDICTION,
    SOURCE_ID,
    OntarioEwrbAdapter,
    OntarioEwrbCaveat,
    OntarioEwrbFetchConfig,
    OntarioEwrbMetricsMapper,
    OntarioEwrbReportMapper,
    OntarioEwrbSubjectMapper,
)
from civix.domains.building_energy_emissions.models import (
    BuildingSubjectKind,
    EmissionsMetricType,
    EnergyMetricType,
    IdentityCertainty,
    MetricFamily,
    MetricValueSource,
    NumericMetricMeasure,
    ReportingPeriodPrecision,
    SourceValueState,
    WaterMetricType,
    build_building_energy_report_key,
    build_building_energy_subject_key,
)

PINNED_NOW = datetime(2026, 5, 3, 12, 0, tzinfo=UTC)
FIXTURES = Path(__file__).parent / "fixtures"
XLSX_FIXTURE = FIXTURES / "ewrb_2024_trimmed.xlsx"


def _xlsx_bytes(path: Path = XLSX_FIXTURE) -> bytes:
    return path.read_bytes()


def _adapter(client: httpx.AsyncClient) -> OntarioEwrbAdapter:
    return OntarioEwrbAdapter(
        fetch_config=OntarioEwrbFetchConfig(client=client, clock=lambda: PINNED_NOW),
    )


def _snapshot() -> SourceSnapshot:
    return SourceSnapshot(
        snapshot_id=SnapshotId("snap-ontario-ewrb"),
        source_id=SOURCE_ID,
        dataset_id=EWRB_DATASET_ID,
        jurisdiction=ONTARIO_JURISDICTION,
        fetched_at=PINNED_NOW,
        record_count=len(_rows()),
        source_url=EWRB_DEFAULT_URL,
        fetch_params={
            "sheet_name": EWRB_DEFAULT_SHEET_NAME,
            "reporting_year": str(EWRB_DEFAULT_REPORTING_YEAR),
        },
    )


def _record(raw: dict[str, Any], source_record_id: str) -> RawRecord:
    return RawRecord(
        snapshot_id=SnapshotId("snap-ontario-ewrb"),
        raw_data=raw,
        source_record_id=source_record_id,
    )


def _rows() -> list[dict[str, str | None]]:
    """Decode fixture rows into the shape the adapter produces, including the
    synthetic `_reporting_year` field the adapter injects on every record."""
    from civix.domains.building_energy_emissions.adapters.sources.ca.ontario_ewrb.schema import (
        EWRB_HEADER_NORMALIZATION,
    )

    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    sheet = workbook[EWRB_DEFAULT_SHEET_NAME]
    iterator = sheet.iter_rows(values_only=True)
    headers_raw = next(iterator)
    headers = [
        EWRB_HEADER_NORMALIZATION.get(str(value).strip()) if value is not None else None
        for value in headers_raw
    ]
    rows: list[dict[str, str | None]] = []

    for row in iterator:
        record: dict[str, str | None] = {}
        for header, value in zip(headers, row, strict=True):
            if header is None:
                continue

            record[header] = _stringify(value)

        record[EWRB_REPORTING_YEAR_FIELD] = str(EWRB_DEFAULT_REPORTING_YEAR)
        rows.append(record)

    workbook.close()

    return rows


def _stringify(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    return str(value)


def _row_record(index: int) -> RawRecord:
    raw = _rows()[index]
    ewrb_id = raw["ewrb_id"]
    year = raw[EWRB_REPORTING_YEAR_FIELD]

    return _record(raw, f"{ewrb_id}:{year}")


def _capture_sequence(
    requests: list[httpx.Request],
    responses: list[httpx.Response],
) -> Callable[[httpx.Request], httpx.Response]:
    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)

        return responses.pop(0)

    return handler


def _empty_workbook_payload() -> bytes:
    workbook = Workbook()
    default = workbook.active

    if default is not None:
        workbook.remove(default)

    sheet = workbook.create_sheet(EWRB_DEFAULT_SHEET_NAME)
    sheet.append(["EWRB_ID", "City"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def _missing_sheet_payload() -> bytes:
    workbook = Workbook()
    default = workbook.active

    if default is not None:
        workbook.remove(default)

    workbook.create_sheet("OtherSheet")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


class TestAdapter:
    async def test_fetches_xlsx_and_emits_one_record_per_row(self) -> None:
        requests: list[httpx.Request] = []

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                side_effect=_capture_sequence(
                    requests,
                    [httpx.Response(200, content=_xlsx_bytes())],
                )
            )

            async with httpx.AsyncClient() as client:
                adapter = _adapter(client)
                result = await adapter.fetch()
                records = [record async for record in result.records]

        assert isinstance(adapter, SourceAdapter)
        assert requests[0].url == EWRB_DEFAULT_URL
        assert result.snapshot.source_id == SOURCE_ID
        assert result.snapshot.dataset_id == EWRB_DATASET_ID
        assert result.snapshot.jurisdiction == ONTARIO_JURISDICTION
        assert result.snapshot.source_url == EWRB_DEFAULT_URL
        assert result.snapshot.record_count == 4
        assert result.snapshot.fetch_params == {
            "sheet_name": EWRB_DEFAULT_SHEET_NAME,
            "reporting_year": str(EWRB_DEFAULT_REPORTING_YEAR),
        }
        assert [record.source_record_id for record in records] == [
            "100001:2024",
            "100002:2024",
            "100003:2024",
            "100004:2024",
        ]
        first = records[0]
        assert first.raw_data["ewrb_id"] == "100001"
        assert first.raw_data["city"] == "Toronto"
        assert first.raw_data["postal_code"] == "M5V"
        assert first.raw_data[EWRB_REPORTING_YEAR_FIELD] == "2024"

    async def test_mixed_city_fixture_includes_toronto_and_non_toronto_rows(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=_xlsx_bytes())
            )

            async with httpx.AsyncClient() as client:
                fetch_result = await _adapter(client).fetch()
                cities = [record.raw_data.get("city") async for record in fetch_result.records]

        assert "Toronto" in cities
        assert any(city != "Toronto" and city is not None for city in cities)

    async def test_empty_workbook_yields_no_records(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=_empty_workbook_payload())
            )

            async with httpx.AsyncClient() as client:
                result = await _adapter(client).fetch()
                records = [record async for record in result.records]

        assert records == []
        assert result.snapshot.record_count == 0

    @pytest.mark.parametrize(
        ("payload_factory", "operation"),
        [
            (lambda: b"not an xlsx", "open-workbook"),
            (_missing_sheet_payload, "open-workbook"),
        ],
    )
    async def test_rejects_malformed_xlsx_payloads(
        self,
        payload_factory: Callable[[], bytes],
        operation: str,
    ) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=payload_factory())
            )

            async with httpx.AsyncClient() as client:
                with pytest.raises(FetchError) as exc_info:
                    await _adapter(client).fetch()

        assert exc_info.value.operation == operation
        assert exc_info.value.dataset_id == EWRB_DATASET_ID

    async def test_http_error_surfaces_as_fetch_error(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(return_value=httpx.Response(503, text="oops"))

            async with httpx.AsyncClient() as client:
                with pytest.raises(FetchError) as exc_info:
                    await _adapter(client).fetch()

        assert exc_info.value.operation == "download"
        assert exc_info.value.dataset_id == EWRB_DATASET_ID


class TestSubjectMapper:
    def test_emits_reporting_account_subject_with_ambiguous_identity(self) -> None:
        result = OntarioEwrbSubjectMapper()(_row_record(0), _snapshot())
        subject = result.record

        assert subject.subject_key == build_building_energy_subject_key(
            SOURCE_ID, EWRB_DATASET_ID, "100001"
        )
        assert subject.subject_kind.value is BuildingSubjectKind.REPORTING_ACCOUNT
        assert subject.identity_certainty.value is IdentityCertainty.AMBIGUOUS
        assert subject.parent_subject_key.value is None

        identifiers = subject.source_subject_identifiers.value or ()
        assert len(identifiers) == 1
        assert identifiers[0].value == "100001"
        assert identifiers[0].identifier_kind is not None
        assert identifiers[0].identifier_kind.code == "ontario-ewrb-id"

    def test_floor_area_marked_redacted_to_distinguish_from_missing(self) -> None:
        subject = OntarioEwrbSubjectMapper()(_row_record(0), _snapshot()).record

        assert subject.floor_area.value is None
        assert subject.floor_area.quality is FieldQuality.REDACTED
        assert subject.floor_area_unit.quality is FieldQuality.REDACTED

    def test_partial_postal_kept_as_published_with_caveat(self) -> None:
        subject = OntarioEwrbSubjectMapper()(_row_record(0), _snapshot()).record

        assert subject.address.quality is FieldQuality.DIRECT
        assert subject.address.value is not None
        assert subject.address.value.postal_code == "M5V"
        assert subject.address.value.locality == "Toronto"
        assert subject.address.value.street is None

        codes = {caveat.code for caveat in (subject.source_caveats.value or ())}
        assert OntarioEwrbCaveat.PARTIAL_POSTAL_DISCLOSURE_FSA_ONLY.value in codes
        assert OntarioEwrbCaveat.SUPPRESSED_TOTAL_METRICS_AND_FLOOR_AREA.value in codes

    def test_dataset_caveats_attached_to_every_subject(self) -> None:
        subject = OntarioEwrbSubjectMapper()(_row_record(2), _snapshot()).record
        caveats = subject.source_caveats.value or ()

        assert all(caveat.taxonomy_id == EWRB_CAVEAT_TAXONOMY_ID for caveat in caveats)
        codes = {caveat.code for caveat in caveats}
        assert OntarioEwrbCaveat.OWNER_REPORTED_NOT_CLEANSED.value in codes
        assert OntarioEwrbCaveat.NRCAN_SOURCE_FACTOR_CHANGE_2023_08_28.value in codes
        assert OntarioEwrbCaveat.OPEN_GOVERNMENT_LICENCE_ONTARIO.value in codes

    def test_missing_ewrb_id_raises_mapper_scoped_error(self) -> None:
        raw = dict(_rows()[0])
        raw["ewrb_id"] = None

        with pytest.raises(MappingError) as exc_info:
            OntarioEwrbSubjectMapper()(_record(raw, "None:2024"), _snapshot())

        assert exc_info.value.source_fields == ("ewrb_id",)


class TestReportMapper:
    def test_report_key_is_deterministic_from_ewrb_id_and_year(self) -> None:
        report = OntarioEwrbReportMapper()(_row_record(0), _snapshot()).record

        assert report.report_key == build_building_energy_report_key(
            SOURCE_ID, EWRB_DATASET_ID, "100001", "2024"
        )
        assert report.subject_key == build_building_energy_subject_key(
            SOURCE_ID, EWRB_DATASET_ID, "100001"
        )
        assert report.reporting_period.value is not None
        assert report.reporting_period.value.precision is TemporalPeriodPrecision.YEAR
        assert report.reporting_period.value.year_value == 2024
        assert report.reporting_period_precision.value is ReportingPeriodPrecision.CALENDAR_YEAR

    def test_data_quality_caveats_flag_checker_not_run(self) -> None:
        report = OntarioEwrbReportMapper()(_row_record(2), _snapshot()).record
        labels = {category.label for category in (report.data_quality_caveats.value or ())}

        assert "Data Quality Checker Not Run" in labels


class TestMetricsMapper:
    def test_complete_row_emits_one_metric_per_disclosed_field_plus_withheld_totals(
        self,
    ) -> None:
        metrics = OntarioEwrbMetricsMapper()(_row_record(0), _snapshot()).record

        # 20 disclosed metric specs + 4 synthetic withheld total metrics.
        assert len(metrics) == 24

        disclosed = [
            metric for metric in metrics if metric.value_state.value is SourceValueState.REPORTED
        ]
        withheld = [
            metric for metric in metrics if metric.value_state.value is SourceValueState.WITHHELD
        ]

        assert len(disclosed) == 20
        assert len(withheld) == 4
        assert all(
            metric.value_source.value is MetricValueSource.SOURCE_PUBLISHED for metric in metrics
        )
        assert all(metric.case_key.quality is FieldQuality.UNMAPPED for metric in metrics)

    def test_withheld_total_metrics_carry_distinct_value_state(self) -> None:
        metrics = OntarioEwrbMetricsMapper()(_row_record(0), _snapshot()).record
        withheld = [
            metric for metric in metrics if metric.value_state.value is SourceValueState.WITHHELD
        ]

        labels = {
            metric.source_metric_label.value.code
            for metric in withheld
            if metric.source_metric_label.value is not None
        }
        assert labels == {
            "total-site-energy-withheld",
            "total-natural-gas-use-withheld",
            "total-ghg-withheld",
            "total-water-use-withheld",
        }
        assert all(metric.measure.value is None for metric in withheld)
        assert all(metric.measure.quality is FieldQuality.REDACTED for metric in withheld)

    def test_post_aug_2023_factor_flag_propagates_to_source_eui_metrics(self) -> None:
        metrics = OntarioEwrbMetricsMapper()(_row_record(0), _snapshot()).record
        source_eui = [
            metric
            for metric in metrics
            if metric.energy_metric_type.value
            in (
                EnergyMetricType.SOURCE_EUI,
                EnergyMetricType.WEATHER_NORMALIZED_SOURCE_EUI,
            )
        ]

        assert source_eui
        for metric in source_eui:
            assert metric.methodology_version.value == METHODOLOGY_VERSION_POST_AUG_2023

        site_eui_only = next(
            metric
            for metric in metrics
            if metric.energy_metric_type.value is EnergyMetricType.SITE_EUI
        )
        # Site EUI is not affected by the source-factor change.
        assert site_eui_only.methodology_version.quality is FieldQuality.UNMAPPED

    def test_pre_aug_2023_flag_propagates_distinct_methodology_version(self) -> None:
        metrics = OntarioEwrbMetricsMapper()(_row_record(1), _snapshot()).record
        source_eui = next(
            metric
            for metric in metrics
            if metric.energy_metric_type.value is EnergyMetricType.SOURCE_EUI
        )

        assert source_eui.methodology_version.value == METHODOLOGY_VERSION_PRE_AUG_2023

    def test_not_available_sentinel_preserves_distinct_value_state(self) -> None:
        metrics = OntarioEwrbMetricsMapper()(_row_record(2), _snapshot()).record
        by_label = {
            metric.source_metric_label.value.code: metric  # type: ignore[union-attr]
            for metric in metrics
            if metric.source_metric_label.value is not None
        }

        electricity = by_label["weather-normalized-electricity-intensity-gj-per-m2"]
        score = by_label["energy-star-score-score"]

        assert electricity.value_state.value is SourceValueState.NOT_AVAILABLE
        assert electricity.measure.value is None
        assert electricity.measure.quality is FieldQuality.NOT_PROVIDED

        assert score.value_state.value is SourceValueState.NOT_AVAILABLE
        assert score.measure.value is None

    def test_disclosed_intensity_carries_typed_metric_type_and_unit(self) -> None:
        metrics = OntarioEwrbMetricsMapper()(_row_record(0), _snapshot()).record

        ghg_intensity = next(
            metric
            for metric in metrics
            if metric.emissions_metric_type.value is EmissionsMetricType.EMISSIONS_INTENSITY
            and metric.unit.value is not None
            and metric.unit.value.code == "kgco2e-per-m2"
        )

        assert ghg_intensity.metric_family is MetricFamily.EMISSIONS
        assert isinstance(ghg_intensity.measure.value, NumericMetricMeasure)
        assert ghg_intensity.measure.value.value == Decimal("18.7")

        water_intensity = next(
            metric
            for metric in metrics
            if metric.water_metric_type.value is WaterMetricType.WATER_USE_INTENSITY
            and metric.source_metric_label.value is not None
            and metric.source_metric_label.value.code == "indoor-water-intensity-m3-per-m2"
        )
        assert water_intensity.metric_family is MetricFamily.WATER

    def test_invalid_decimal_raises_mapping_error_with_source_field(self) -> None:
        raw = dict(_rows()[0])
        raw["site_eui_gj_per_m2"] = "1.32abc"

        with pytest.raises(MappingError) as exc_info:
            OntarioEwrbMetricsMapper()(_record(raw, "100001:2024"), _snapshot())

        assert exc_info.value.source_fields == ("site_eui_gj_per_m2",)


class TestPipelineDrift:
    async def test_fixture_drift_clean(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=_xlsx_bytes())
            )

            async with httpx.AsyncClient() as client:
                pipeline_result = await run(_adapter(client), OntarioEwrbSubjectMapper())
                schema_obs = SchemaObserver(spec=EWRB_RAW_SCHEMA)
                taxonomy_obs = TaxonomyObserver(specs=EWRB_TAXONOMIES)
                observed = attach_observers(pipeline_result, [schema_obs, taxonomy_obs])
                consumed = [record async for record in observed.records]

        assert len(consumed) == 4
        assert schema_obs.finalize(pipeline_result.snapshot).findings == ()
        assert taxonomy_obs.finalize(pipeline_result.snapshot).findings == ()

    async def test_unknown_property_type_surfaces_taxonomy_drift(self) -> None:
        workbook = load_workbook(XLSX_FIXTURE, data_only=True)
        sheet = workbook[EWRB_DEFAULT_SHEET_NAME]
        headers = [str(value) for value in next(sheet.iter_rows(max_row=1, values_only=True))]
        column = headers.index("PrimPropTypCalc") + 1
        sheet.cell(row=2, column=column, value="Aquatic Centre")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=output.getvalue())
            )

            async with httpx.AsyncClient() as client:
                fetch_result = await _adapter(client).fetch()
                taxonomy_obs = TaxonomyObserver(specs=EWRB_TAXONOMIES)
                async for record in fetch_result.records:
                    taxonomy_obs.observe(record)

        report = taxonomy_obs.finalize(fetch_result.snapshot)

        assert any(
            finding.kind is TaxonomyDriftKind.UNRECOGNIZED_VALUE
            and finding.taxonomy_id == "ontario-ewrb-primary-property-type"
            for finding in report.findings
        )

    async def test_renamed_required_column_surfaces_schema_drift(self) -> None:
        workbook = load_workbook(XLSX_FIXTURE, data_only=True)
        sheet = workbook[EWRB_DEFAULT_SHEET_NAME]
        sheet.cell(row=1, column=1, value="EWRB_Reporting_ID")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(EWRB_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=output.getvalue())
            )

            async with httpx.AsyncClient() as client:
                fetch_result = await _adapter(client).fetch()
                schema_obs = SchemaObserver(spec=EWRB_RAW_SCHEMA)
                async for record in fetch_result.records:
                    schema_obs.observe(record)

        report = schema_obs.finalize(fetch_result.snapshot)

        assert any(
            finding.kind is SchemaDriftKind.MISSING_FIELD and finding.field_name == "ewrb_id"
            for finding in report.findings
        )


def test_source_metadata_preserves_scope_and_layout() -> None:
    assert SOURCE_ID == "ontario-open-data"
    assert "Ontario" in EWRB_SOURCE_SCOPE
    assert EWRB_DEFAULT_SHEET_NAME == "Sheet1"
    assert "ewrb_id" in {spec.name for spec in EWRB_RAW_SCHEMA.fields}
    assert EWRB_REPORTING_YEAR_FIELD in {spec.name for spec in EWRB_RAW_SCHEMA.fields}
