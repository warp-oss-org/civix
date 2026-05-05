"""Tests for the NYC LL97 covered-buildings source slice."""

from __future__ import annotations

from collections.abc import Callable
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
import pytest
import respx
from openpyxl import Workbook, load_workbook

from civix.core.drift import (
    SchemaDriftKind,
    SchemaObserver,
    TaxonomyDriftKind,
    TaxonomyObserver,
)
from civix.core.identity.models.identifiers import SnapshotId
from civix.core.mapping.errors import MappingError
from civix.core.pipeline import attach_observers, run
from civix.core.ports.errors import FetchError
from civix.core.ports.models.adapter import SourceAdapter
from civix.core.quality.models.fields import FieldQuality
from civix.core.snapshots.models.snapshot import RawRecord, SourceSnapshot
from civix.core.temporal import TemporalPeriodPrecision
from civix.domains.building_energy_emissions.adapters.sources.us.nyc_ll97 import (
    LL97_CAVEAT_TAXONOMY_ID,
    LL97_DATASET_ID,
    LL97_DEFAULT_FILING_YEAR,
    LL97_DEFAULT_URL,
    LL97_RAW_SCHEMA,
    LL97_SHEET_NAME,
    LL97_SOURCE_SCOPE,
    LL97_TAXONOMIES,
    NYC_JURISDICTION,
    SOURCE_ID,
    NycLl97Adapter,
    NycLl97CaseMapper,
    NycLl97Caveat,
    NycLl97FetchConfig,
    NycLl97SubjectMapper,
)
from civix.domains.building_energy_emissions.models import (
    BuildingSubjectKind,
    ComplianceLifecycleStatus,
    IdentityCertainty,
    build_building_compliance_case_key,
    build_building_energy_subject_key,
)

PINNED_NOW = datetime(2026, 5, 3, 12, 0, tzinfo=UTC)
FIXTURES = Path(__file__).parent / "fixtures"
XLSX_FIXTURE = FIXTURES / "cbl26_trimmed.xlsx"


def _xlsx_bytes(path: Path = XLSX_FIXTURE) -> bytes:
    return path.read_bytes()


def _adapter(client: httpx.AsyncClient) -> NycLl97Adapter:
    return NycLl97Adapter(
        fetch_config=NycLl97FetchConfig(client=client, clock=lambda: PINNED_NOW),
    )


def _snapshot() -> SourceSnapshot:
    return SourceSnapshot(
        snapshot_id=SnapshotId("snap-nyc-ll97"),
        source_id=SOURCE_ID,
        dataset_id=LL97_DATASET_ID,
        jurisdiction=NYC_JURISDICTION,
        fetched_at=PINNED_NOW,
        record_count=len(_rows()),
        source_url=LL97_DEFAULT_URL,
        fetch_params={
            "sheet_name": LL97_SHEET_NAME,
            "filing_year": str(LL97_DEFAULT_FILING_YEAR),
        },
    )


def _record(raw: dict[str, Any], source_record_id: str) -> RawRecord:
    return RawRecord(
        snapshot_id=SnapshotId("snap-nyc-ll97"),
        raw_data=raw,
        source_record_id=source_record_id,
    )


def _rows() -> list[dict[str, str | None]]:
    """Decode the fixture rows into the canonical snake_case shape the
    adapter produces. Mirrors the adapter's normalization+stringification
    so mapper tests can drive records without re-hitting the workbook."""
    workbook = load_workbook(XLSX_FIXTURE, data_only=True)
    sheet = workbook[LL97_SHEET_NAME]
    iterator = sheet.iter_rows(values_only=True)
    headers_raw = next(iterator)
    headers = [_normalize(str(value)) for value in headers_raw]
    rows: list[dict[str, str | None]] = []

    for row in iterator:
        rows.append({header: _stringify(value) for header, value in zip(headers, row, strict=True)})

    workbook.close()

    return rows


def _normalize(value: str) -> str:
    import re

    stripped = value.strip().lower()
    no_parens = re.sub(r"\([^)]*\)", "", stripped)
    underscored = re.sub(r"[^a-z0-9]+", "_", no_parens)

    return underscored.strip("_")


def _stringify(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    return str(value)


def _row_record(index: int) -> RawRecord:
    raw = _rows()[index]
    bbl = raw["bbl"]
    bin_ = raw["bin"]

    return _record(raw, f"{bbl}:{bin_}")


def _capture_sequence(
    requests: list[httpx.Request],
    responses: list[httpx.Response],
) -> Callable[[httpx.Request], httpx.Response]:
    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)

        return responses.pop(0)

    return handler


def _empty_workbook_payload() -> bytes:
    workbook = Workbook()
    default = workbook.active

    if default is not None:
        workbook.remove(default)

    sheet = workbook.create_sheet(LL97_SHEET_NAME)
    sheet.append([str(spec.name) for spec in LL97_RAW_SCHEMA.fields])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def _missing_sheet_payload() -> bytes:
    workbook = Workbook()
    default = workbook.active

    if default is not None:
        workbook.remove(default)

    workbook.create_sheet("Some_Other_Sheet")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


class TestAdapter:
    async def test_fetches_xlsx_and_emits_one_record_per_covered_row(self) -> None:
        requests: list[httpx.Request] = []

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(
                side_effect=_capture_sequence(
                    requests,
                    [httpx.Response(200, content=_xlsx_bytes())],
                )
            )

            async with httpx.AsyncClient() as client:
                adapter = _adapter(client)
                result = await adapter.fetch()
                records = [record async for record in result.records]

        assert isinstance(adapter, SourceAdapter)
        assert requests[0].url == LL97_DEFAULT_URL
        assert result.snapshot.source_id == SOURCE_ID
        assert result.snapshot.dataset_id == LL97_DATASET_ID
        assert result.snapshot.jurisdiction == NYC_JURISDICTION
        assert result.snapshot.source_url == LL97_DEFAULT_URL
        assert result.snapshot.record_count == 5
        assert result.snapshot.fetch_params == {
            "sheet_name": LL97_SHEET_NAME,
            "filing_year": str(LL97_DEFAULT_FILING_YEAR),
        }

        assert [record.source_record_id for record in records] == [
            "1009990001:1011223",
            "1009990002:1011224",
            "2034560020:2050010",
            "2034560020:2050011",
            "3001000004:3001500",
        ]
        first = records[0]

        assert first.raw_data["bbl"] == "1009990001"
        assert first.raw_data["on_ll97_cbl"] == "Y"
        assert first.raw_data["ll97_compliance_pathway"] == "2"

    async def test_empty_workbook_yields_no_records(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=_empty_workbook_payload())
            )

            async with httpx.AsyncClient() as client:
                result = await _adapter(client).fetch()
                records = [record async for record in result.records]

        assert records == []
        assert result.snapshot.record_count == 0

    @pytest.mark.parametrize(
        ("payload_factory", "operation"),
        [
            (lambda: b"not an xlsx", "open-workbook"),
            (_missing_sheet_payload, "open-workbook"),
        ],
    )
    async def test_rejects_malformed_xlsx_payloads(
        self,
        payload_factory: Callable[[], bytes],
        operation: str,
    ) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=payload_factory())
            )

            async with httpx.AsyncClient() as client:
                with pytest.raises(FetchError) as exc_info:
                    await _adapter(client).fetch()

        assert exc_info.value.operation == operation
        assert exc_info.value.dataset_id == LL97_DATASET_ID

    async def test_http_error_surfaces_as_fetch_error(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(return_value=httpx.Response(503, text="oops"))

            async with httpx.AsyncClient() as client:
                with pytest.raises(FetchError) as exc_info:
                    await _adapter(client).fetch()

        assert exc_info.value.operation == "download"
        assert exc_info.value.dataset_id == LL97_DATASET_ID


class TestSubjectMapper:
    def test_emits_bin_centered_subject_with_bbl_preserved_separately(self) -> None:
        result = NycLl97SubjectMapper()(_row_record(0), _snapshot())
        subject = result.record

        assert subject.subject_key == build_building_energy_subject_key(
            SOURCE_ID, LL97_DATASET_ID, "1011223"
        )

        assert subject.subject_kind.value is BuildingSubjectKind.BUILDING
        assert subject.identity_certainty.value is IdentityCertainty.STABLE_CROSS_YEAR
        assert subject.parent_subject_key.quality is FieldQuality.UNMAPPED
        assert subject.parent_subject_key.value is None

        identifiers = subject.source_subject_identifiers.value or ()
        identifier_kinds = [
            identifier.identifier_kind.code
            for identifier in identifiers
            if identifier.identifier_kind is not None
        ]
        identifier_values = [identifier.value for identifier in identifiers]

        assert identifier_kinds == ["dob-bin", "dof-bbl"]
        assert identifier_values == ["1011223", "1009990001"]

        assert subject.address.quality is FieldQuality.DIRECT
        assert subject.address.value is not None
        assert subject.address.value.street == "123 First Avenue"
        assert subject.address.value.postal_code == "10003"

    def test_multi_building_lot_emits_distinct_subject_per_bin_with_shared_bbl(self) -> None:
        first = NycLl97SubjectMapper()(_row_record(2), _snapshot()).record
        second = NycLl97SubjectMapper()(_row_record(3), _snapshot()).record

        assert first.subject_key != second.subject_key
        assert first.subject_key == build_building_energy_subject_key(
            SOURCE_ID, LL97_DATASET_ID, "2050010"
        )

        assert second.subject_key == build_building_energy_subject_key(
            SOURCE_ID, LL97_DATASET_ID, "2050011"
        )

        first_bbls = [
            identifier.value
            for identifier in (first.source_subject_identifiers.value or ())
            if identifier.identifier_kind is not None
            and identifier.identifier_kind.code == "dof-bbl"
        ]
        second_bbls = [
            identifier.value
            for identifier in (second.source_subject_identifiers.value or ())
            if identifier.identifier_kind is not None
            and identifier.identifier_kind.code == "dof-bbl"
        ]

        assert first_bbls == ["2034560020"]
        assert second_bbls == ["2034560020"]

        assert first.parent_subject_key.value is None
        assert second.parent_subject_key.value is None

    def test_dataset_caveats_attached_to_every_subject(self) -> None:
        subject = NycLl97SubjectMapper()(_row_record(0), _snapshot()).record
        caveats = subject.source_caveats.value or ()
        codes = {caveat.code for caveat in caveats}

        assert NycLl97Caveat.PRELIMINARY_LIST_REFERENCE_ONLY.value in codes
        assert NycLl97Caveat.OWNER_DISPUTABLE.value in codes
        assert all(caveat.taxonomy_id == LL97_CAVEAT_TAXONOMY_ID for caveat in caveats)

    def test_missing_bin_raises_mapper_scoped_error(self) -> None:
        raw = dict(_rows()[0])
        raw["bin"] = None

        with pytest.raises(MappingError) as exc_info:
            NycLl97SubjectMapper()(_record(raw, "1009990001:None"), _snapshot())

        assert exc_info.value.source_fields == ("bin",)


class TestCaseMapper:
    def test_full_case_emits_pathway_and_filing_year_and_leaves_lifecycle_unset(self) -> None:
        result = NycLl97CaseMapper()(_row_record(0), _snapshot())
        case = result.record

        assert case.case_key == build_building_compliance_case_key(
            SOURCE_ID,
            LL97_DATASET_ID,
            "1009990001:1011223",
            f"filing-year-{LL97_DEFAULT_FILING_YEAR}",
        )

        assert case.subject_key == build_building_energy_subject_key(
            SOURCE_ID, LL97_DATASET_ID, "1011223"
        )

        assert case.related_report_key.quality is FieldQuality.UNMAPPED
        assert case.related_report_key.value is None

        assert case.covered_building_status.value is ComplianceLifecycleStatus.COVERED
        assert case.source_covered_status.value is not None
        assert case.source_covered_status.value.code == "y"

        assert case.compliance_pathway.quality is FieldQuality.STANDARDIZED
        pathway = case.compliance_pathway.value

        assert pathway is not None
        assert pathway.code == "pathway-2"
        assert pathway.taxonomy_id == "nyc-ll97-compliance-pathway"

        assert case.filing_period.value is not None
        assert case.filing_period.value.precision is TemporalPeriodPrecision.YEAR
        assert case.filing_period.value.year_value == LL97_DEFAULT_FILING_YEAR

        for unset in (
            case.compliance_status,
            case.source_compliance_status,
            case.exemption_status,
            case.extension_status,
            case.dispute_status,
            case.penalty_amount,
            case.penalty_currency,
            case.penalty_status,
            case.emissions_limit_metric_key,
            case.final_emissions_metric_key,
            case.excess_emissions_metric_key,
            case.covered_period,
        ):
            assert unset.quality is FieldQuality.UNMAPPED
            assert unset.value is None

    def test_minimum_valid_covered_case_marks_pathway_not_provided(self) -> None:
        case = NycLl97CaseMapper()(_row_record(1), _snapshot()).record

        assert case.covered_building_status.value is ComplianceLifecycleStatus.COVERED
        assert case.compliance_pathway.quality is FieldQuality.NOT_PROVIDED
        assert case.compliance_pathway.value is None
        assert case.compliance_pathway.source_fields == ("ll97_compliance_pathway",)

    def test_not_covered_row_marks_pathway_unmapped_not_not_provided(self) -> None:
        case = NycLl97CaseMapper()(_row_record(4), _snapshot()).record

        assert case.covered_building_status.value is ComplianceLifecycleStatus.NOT_COVERED
        assert case.source_covered_status.value is not None
        assert case.source_covered_status.value.code == "n"
        assert case.compliance_pathway.quality is FieldQuality.UNMAPPED
        assert case.compliance_pathway.value is None
        assert case.compliance_pathway.source_fields == ()

    def test_multi_building_lot_emits_one_case_per_bin(self) -> None:
        first = NycLl97CaseMapper()(_row_record(2), _snapshot()).record
        second = NycLl97CaseMapper()(_row_record(3), _snapshot()).record

        assert first.case_key != second.case_key
        assert first.subject_key != second.subject_key
        assert first.compliance_pathway.value is not None
        assert second.compliance_pathway.value is not None
        assert first.compliance_pathway.value.code == "pathway-1"
        assert second.compliance_pathway.value.code == "pathway-1"

    def test_case_carries_dataset_caveats_including_pathway_without_limit(self) -> None:
        case = NycLl97CaseMapper()(_row_record(0), _snapshot()).record
        codes = {caveat.code for caveat in (case.source_caveats.value or ())}

        assert NycLl97Caveat.PATHWAY_CODE_PUBLISHED_WITHOUT_NUMERIC_LIMIT.value in codes
        assert NycLl97Caveat.PRELIMINARY_LIST_REFERENCE_ONLY.value in codes

    def test_unrecognized_covered_value_raises_mapping_error(self) -> None:
        raw = dict(_rows()[0])
        raw["on_ll97_cbl"] = "maybe"

        with pytest.raises(MappingError) as exc_info:
            NycLl97CaseMapper()(_record(raw, "1009990001:1011223"), _snapshot())

        assert exc_info.value.source_fields == ("on_ll97_cbl",)

    def test_missing_filing_year_in_snapshot_params_raises(self) -> None:
        snapshot = SourceSnapshot(
            snapshot_id=SnapshotId("snap-nyc-ll97-no-year"),
            source_id=SOURCE_ID,
            dataset_id=LL97_DATASET_ID,
            jurisdiction=NYC_JURISDICTION,
            fetched_at=PINNED_NOW,
            record_count=1,
            source_url=LL97_DEFAULT_URL,
        )

        with pytest.raises(MappingError) as exc_info:
            NycLl97CaseMapper()(_row_record(0), snapshot)

        assert "filing_year" in str(exc_info.value)


class TestPipelineDrift:
    async def test_fixture_drift_clean(self) -> None:
        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=_xlsx_bytes())
            )

            async with httpx.AsyncClient() as client:
                pipeline_result = await run(_adapter(client), NycLl97SubjectMapper())
                schema_obs = SchemaObserver(spec=LL97_RAW_SCHEMA)
                taxonomy_obs = TaxonomyObserver(specs=LL97_TAXONOMIES)
                observed = attach_observers(pipeline_result, [schema_obs, taxonomy_obs])
                consumed = [record async for record in observed.records]

        assert len(consumed) == 5
        assert schema_obs.finalize(pipeline_result.snapshot).findings == ()
        assert taxonomy_obs.finalize(pipeline_result.snapshot).findings == ()

    async def test_unknown_pathway_value_surfaces_taxonomy_drift(self) -> None:
        workbook = load_workbook(XLSX_FIXTURE, data_only=True)
        sheet = workbook[LL97_SHEET_NAME]
        headers = [str(value) for value in next(sheet.iter_rows(max_row=1, values_only=True))]
        column = headers.index("LL97 Compliance Pathway") + 1
        sheet.cell(row=2, column=column, value="9")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=output.getvalue())
            )

            async with httpx.AsyncClient() as client:
                fetch_result = await _adapter(client).fetch()
                taxonomy_obs = TaxonomyObserver(specs=LL97_TAXONOMIES)
                async for record in fetch_result.records:
                    taxonomy_obs.observe(record)

        report = taxonomy_obs.finalize(fetch_result.snapshot)

        assert any(
            finding.kind is TaxonomyDriftKind.UNRECOGNIZED_VALUE
            and finding.taxonomy_id == "nyc-ll97-compliance-pathway"
            for finding in report.findings
        )

    async def test_renamed_required_column_surfaces_schema_drift(self) -> None:
        workbook = load_workbook(XLSX_FIXTURE, data_only=True)
        sheet = workbook[LL97_SHEET_NAME]
        sheet.cell(row=1, column=1, value="Borough Block Lot Number")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        async with respx.mock(assert_all_called=True) as respx_mock:
            respx_mock.get(LL97_DEFAULT_URL).mock(
                return_value=httpx.Response(200, content=output.getvalue())
            )

            async with httpx.AsyncClient() as client:
                fetch_result = await _adapter(client).fetch()
                schema_obs = SchemaObserver(spec=LL97_RAW_SCHEMA)
                async for record in fetch_result.records:
                    schema_obs.observe(record)

        report = schema_obs.finalize(fetch_result.snapshot)

        assert any(
            finding.kind is SchemaDriftKind.MISSING_FIELD and finding.field_name == "bbl"
            for finding in report.findings
        )


def test_source_metadata_preserves_scope_and_layout() -> None:
    assert SOURCE_ID == "nyc-dob"
    assert LL97_DATASET_ID == "ll97-covered-buildings-list"
    assert "Local Law 97" in LL97_SOURCE_SCOPE
    assert LL97_SHEET_NAME == "Sustainability_CBL"
    assert "bbl" in {spec.name for spec in LL97_RAW_SCHEMA.fields}
    assert "bin" in {spec.name for spec in LL97_RAW_SCHEMA.fields}
