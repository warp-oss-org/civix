"""NYC LL97 Covered Buildings List source adapter.

DOB publishes the LL97 CBL as an Excel workbook (`cbl{yy}.xlsx`) under
`nyc.gov/assets/buildings/excel/`. The adapter downloads the workbook,
opens the `Sustainability_CBL` sheet, normalizes the inconsistently
spaced header row into canonical snake_case field names, and yields one
`RawRecord` per BIN/BBL row.

There is no shared xlsx infrastructure under `infra/sources/` yet; the
fetch and parse code lives here until a second xlsx source justifies an
extraction.
"""

from __future__ import annotations

import re
from collections.abc import AsyncIterable
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Final

import httpx
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from civix.core.identity.models.identifiers import (
    DatasetId,
    Jurisdiction,
    SnapshotId,
    SourceId,
)
from civix.core.ports.errors import FetchError
from civix.core.ports.models.adapter import FetchResult
from civix.core.snapshots.models.snapshot import RawRecord, SourceSnapshot
from civix.core.temporal import Clock, utc_now
from civix.domains.building_energy_emissions.adapters.sources.us.nyc_ll97.schema import (
    LL97_HEADER_NORMALIZATION,
)

SOURCE_ID: Final[SourceId] = SourceId("nyc-dob")
LL97_DATASET_ID: Final[DatasetId] = DatasetId("ll97-covered-buildings-list")
NYC_JURISDICTION: Final[Jurisdiction] = Jurisdiction(
    country="US", region="NY", locality="New York City"
)
LL97_SHEET_NAME: Final[str] = "Sustainability_CBL"
LL97_DEFAULT_FILING_YEAR: Final[int] = 2026
LL97_DEFAULT_BASE_URL: Final[str] = "https://www.nyc.gov/assets/buildings/excel/"
LL97_DEFAULT_FILENAME: Final[str] = "cbl26.xlsx"
LL97_DEFAULT_URL: Final[str] = f"{LL97_DEFAULT_BASE_URL}{LL97_DEFAULT_FILENAME}"
LL97_SOURCE_SCOPE: Final[str] = (
    "NYC Department of Buildings Sustainability Covered Buildings List, the "
    "annually republished list of buildings covered by Local Law 97 with "
    "compliance pathway codes and tax-lot context."
)
LL97_SOURCE_RECORD_ID_FIELDS: Final[tuple[str, str]] = ("bbl", "bin")


@dataclass(frozen=True, slots=True)
class NycLl97FetchConfig:
    """Runtime fetch options for one NYC LL97 CBL snapshot."""

    client: httpx.AsyncClient
    clock: Clock = field(default=utc_now)
    url: str = LL97_DEFAULT_URL
    sheet_name: str = LL97_SHEET_NAME
    filing_year: int = LL97_DEFAULT_FILING_YEAR

    def __post_init__(self) -> None:
        if not self.url:
            raise ValueError("url must be a non-empty string")

        if not self.sheet_name:
            raise ValueError("sheet_name must be a non-empty string")

        if self.filing_year <= 0:
            raise ValueError("filing_year must be a positive integer")


@dataclass(frozen=True, slots=True)
class NycLl97Adapter:
    """Fetches NYC LL97 CBL rows from the DOB-published xlsx file."""

    fetch_config: NycLl97FetchConfig

    @property
    def source_id(self) -> SourceId:
        return SOURCE_ID

    @property
    def dataset_id(self) -> DatasetId:
        return LL97_DATASET_ID

    @property
    def jurisdiction(self) -> Jurisdiction:
        return NYC_JURISDICTION

    async def fetch(self) -> FetchResult:
        body = await _download(self.fetch_config)
        record_count = _count_data_rows(body, self.fetch_config.sheet_name)
        fetched_at = self.fetch_config.clock()
        snapshot_id = SnapshotId(f"{SOURCE_ID}:{LL97_DATASET_ID}:{fetched_at.isoformat()}")
        snapshot = SourceSnapshot(
            snapshot_id=snapshot_id,
            source_id=SOURCE_ID,
            dataset_id=LL97_DATASET_ID,
            jurisdiction=NYC_JURISDICTION,
            fetched_at=fetched_at,
            record_count=record_count,
            source_url=self.fetch_config.url,
            fetch_params={
                "sheet_name": self.fetch_config.sheet_name,
                "filing_year": str(self.fetch_config.filing_year),
            },
        )

        return FetchResult(
            snapshot=snapshot,
            records=_stream_records(
                body=body,
                sheet_name=self.fetch_config.sheet_name,
                snapshot_id=snapshot_id,
            ),
        )


async def _download(config: NycLl97FetchConfig) -> bytes:
    try:
        response = await config.client.get(config.url)
        response.raise_for_status()
    except httpx.HTTPError as e:
        raise FetchError(
            f"failed to download CBL from {config.url}",
            source_id=SOURCE_ID,
            dataset_id=LL97_DATASET_ID,
            operation="download",
        ) from e

    return response.content


def _count_data_rows(body: bytes, sheet_name: str) -> int:
    workbook = _load_workbook(body)
    try:
        worksheet = _get_sheet(workbook, sheet_name)
        rows_iter = worksheet.iter_rows(values_only=True)
        next(rows_iter, None)  # consume header
        count = sum(1 for _ in rows_iter)
    finally:
        workbook.close()

    return count


async def _stream_records(
    *,
    body: bytes,
    sheet_name: str,
    snapshot_id: SnapshotId,
) -> AsyncIterable[RawRecord]:
    workbook = _load_workbook(body)
    try:
        worksheet = _get_sheet(workbook, sheet_name)
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)

        if header_row is None:
            return

        field_names = _resolve_field_names(header_row)

        for row in rows_iter:
            yield _build_record(
                snapshot_id=snapshot_id,
                field_names=field_names,
                row=row,
            )
    finally:
        workbook.close()


def _build_record(
    *,
    snapshot_id: SnapshotId,
    field_names: tuple[str | None, ...],
    row: tuple[Any, ...],
) -> RawRecord:
    raw_data: dict[str, str | None] = {}

    for column_index, name in enumerate(field_names):
        if name is None:
            continue

        raw_data[name] = _cell_to_string(row[column_index] if column_index < len(row) else None)

    bbl = raw_data.get("bbl")
    bin_ = raw_data.get("bin")
    source_record_id = f"{bbl}:{bin_}" if bbl and bin_ else None

    try:
        return RawRecord(
            snapshot_id=snapshot_id,
            raw_data=raw_data,
            source_record_id=source_record_id,
        )
    except ValidationError as e:
        raise FetchError(
            "raw record failed validation",
            source_id=SOURCE_ID,
            dataset_id=LL97_DATASET_ID,
            operation="stream-records",
        ) from e


def _resolve_field_names(header_row: tuple[Any, ...]) -> tuple[str | None, ...]:
    return tuple(_lookup_field_name(cell) for cell in header_row)


def _lookup_field_name(cell: object) -> str | None:
    if cell is None:
        return None

    normalized = _normalize_header(str(cell))

    return LL97_HEADER_NORMALIZATION.get(normalized)


_PARENTHETICAL = re.compile(r"\([^)]*\)")
_NON_ALNUM_RUN = re.compile(r"[^a-z0-9]+")


def _normalize_header(header: str) -> str:
    stripped = header.strip().lower()
    without_parenthetical = _PARENTHETICAL.sub("", stripped)
    underscored = _NON_ALNUM_RUN.sub("_", without_parenthetical)

    return underscored.strip("_")


def _cell_to_string(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    return str(value)


def _load_workbook(body: bytes) -> Workbook:
    try:
        return openpyxl.load_workbook(BytesIO(body), read_only=True, data_only=True)
    except Exception as e:
        raise FetchError(
            "failed to open CBL workbook",
            source_id=SOURCE_ID,
            dataset_id=LL97_DATASET_ID,
            operation="open-workbook",
        ) from e


def _get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise FetchError(
            f"workbook is missing required sheet {sheet_name!r}",
            source_id=SOURCE_ID,
            dataset_id=LL97_DATASET_ID,
            operation="open-workbook",
        )

    return workbook[sheet_name]
