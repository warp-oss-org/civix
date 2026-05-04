"""Ontario EWRB (Energy and Water Reporting for Buildings) source adapter.

Ontario publishes the EWRB dataset as a yearly XLSX resource on
`data.ontario.ca`. The adapter downloads one yearly resource, opens its
data sheet, normalizes the published header row into canonical
snake_case field names, and yields one `RawRecord` per building/property
row.

There is no shared xlsx infrastructure under `infra/sources/` yet; the
fetch and parse code lives here and intentionally mirrors the
pattern used by the NYC LL97 CBL adapter.
"""

from __future__ import annotations

from collections.abc import AsyncIterable
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Final

import httpx
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from civix.core.identity.models.identifiers import (
    DatasetId,
    Jurisdiction,
    SnapshotId,
    SourceId,
)
from civix.core.ports.errors import FetchError
from civix.core.ports.models.adapter import FetchResult
from civix.core.snapshots.models.snapshot import RawRecord, SourceSnapshot
from civix.core.temporal import Clock, utc_now
from civix.domains.building_energy_emissions.adapters.sources.ca.ontario_ewrb.schema import (
    EWRB_HEADER_NORMALIZATION,
    EWRB_REPORTING_YEAR_FIELD,
)

SOURCE_ID: Final[SourceId] = SourceId("ontario-open-data")
EWRB_DATASET_ID: Final[DatasetId] = DatasetId(
    "energy-and-water-usage-of-large-buildings-in-ontario"
)
ONTARIO_JURISDICTION: Final[Jurisdiction] = Jurisdiction(country="CA", region="ON")
EWRB_DEFAULT_SHEET_NAME: Final[str] = "Sheet1"
EWRB_DEFAULT_REPORTING_YEAR: Final[int] = 2024
EWRB_DEFAULT_BASE_URL: Final[str] = "https://files.ontario.ca/opendata/"
EWRB_DEFAULT_FILENAME: Final[str] = "ewrb-2024.xlsx"
EWRB_DEFAULT_URL: Final[str] = f"{EWRB_DEFAULT_BASE_URL}{EWRB_DEFAULT_FILENAME}"
EWRB_SOURCE_SCOPE: Final[str] = (
    "Ontario Energy and Water Reporting for Buildings (EWRB) annual "
    "building/property-level intensity disclosures, published by the "
    "Government of Ontario via data.ontario.ca."
)
EWRB_SOURCE_RECORD_ID_FIELDS: Final[tuple[str, str]] = ("ewrb_id", EWRB_REPORTING_YEAR_FIELD)


@dataclass(frozen=True, slots=True)
class OntarioEwrbFetchConfig:
    """Runtime fetch options for one Ontario EWRB yearly snapshot."""

    client: httpx.AsyncClient
    clock: Clock = field(default=utc_now)
    url: str = EWRB_DEFAULT_URL
    sheet_name: str = EWRB_DEFAULT_SHEET_NAME
    reporting_year: int = EWRB_DEFAULT_REPORTING_YEAR

    def __post_init__(self) -> None:
        if not self.url:
            raise ValueError("url must be a non-empty string")

        if not self.sheet_name:
            raise ValueError("sheet_name must be a non-empty string")

        if self.reporting_year <= 0:
            raise ValueError("reporting_year must be a positive integer")


@dataclass(frozen=True, slots=True)
class OntarioEwrbAdapter:
    """Fetches one yearly Ontario EWRB workbook from data.ontario.ca."""

    fetch_config: OntarioEwrbFetchConfig

    @property
    def source_id(self) -> SourceId:
        return SOURCE_ID

    @property
    def dataset_id(self) -> DatasetId:
        return EWRB_DATASET_ID

    @property
    def jurisdiction(self) -> Jurisdiction:
        return ONTARIO_JURISDICTION

    async def fetch(self) -> FetchResult:
        body = await _download(self.fetch_config)
        field_names, data_rows = _read_workbook_rows(body, self.fetch_config.sheet_name)
        fetched_at = self.fetch_config.clock()
        snapshot_id = SnapshotId(f"{SOURCE_ID}:{EWRB_DATASET_ID}:{fetched_at.isoformat()}")
        snapshot = SourceSnapshot(
            snapshot_id=snapshot_id,
            source_id=SOURCE_ID,
            dataset_id=EWRB_DATASET_ID,
            jurisdiction=ONTARIO_JURISDICTION,
            fetched_at=fetched_at,
            record_count=len(data_rows),
            source_url=self.fetch_config.url,
            fetch_params={
                "sheet_name": self.fetch_config.sheet_name,
                "reporting_year": str(self.fetch_config.reporting_year),
            },
        )

        return FetchResult(
            snapshot=snapshot,
            records=_stream_records(
                field_names=field_names,
                data_rows=data_rows,
                reporting_year=self.fetch_config.reporting_year,
                snapshot_id=snapshot_id,
            ),
        )


async def _download(config: OntarioEwrbFetchConfig) -> bytes:
    try:
        response = await config.client.get(config.url)
        response.raise_for_status()
    except httpx.HTTPError as e:
        raise FetchError(
            f"failed to download EWRB workbook from {config.url}",
            source_id=SOURCE_ID,
            dataset_id=EWRB_DATASET_ID,
            operation="download",
        ) from e

    return response.content


def _read_workbook_rows(
    body: bytes, sheet_name: str
) -> tuple[tuple[str | None, ...], tuple[tuple[Any, ...], ...]]:
    workbook = _load_workbook(body)
    try:
        worksheet = _get_sheet(workbook, sheet_name)
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)

        if header_row is None:
            return (), ()

        field_names = _resolve_field_names(header_row)
        data_rows = tuple(rows_iter)
    finally:
        workbook.close()

    return field_names, data_rows


async def _stream_records(
    *,
    field_names: tuple[str | None, ...],
    data_rows: tuple[tuple[Any, ...], ...],
    reporting_year: int,
    snapshot_id: SnapshotId,
) -> AsyncIterable[RawRecord]:
    reporting_year_text = str(reporting_year)

    for row in data_rows:
        yield _build_record(
            snapshot_id=snapshot_id,
            field_names=field_names,
            row=row,
            reporting_year_text=reporting_year_text,
        )


def _build_record(
    *,
    snapshot_id: SnapshotId,
    field_names: tuple[str | None, ...],
    row: tuple[Any, ...],
    reporting_year_text: str,
) -> RawRecord:
    raw_data: dict[str, str | None] = {}

    for column_index, name in enumerate(field_names):
        if name is None:
            continue

        raw_data[name] = _cell_to_string(row[column_index] if column_index < len(row) else None)

    raw_data[EWRB_REPORTING_YEAR_FIELD] = reporting_year_text

    ewrb_id = raw_data.get("ewrb_id")
    source_record_id = f"{ewrb_id}:{reporting_year_text}" if ewrb_id else None

    try:
        return RawRecord(
            snapshot_id=snapshot_id,
            raw_data=raw_data,
            source_record_id=source_record_id,
        )
    except ValidationError as e:
        raise FetchError(
            "raw record failed validation",
            source_id=SOURCE_ID,
            dataset_id=EWRB_DATASET_ID,
            operation="stream-records",
        ) from e


def _resolve_field_names(header_row: tuple[Any, ...]) -> tuple[str | None, ...]:
    return tuple(_lookup_field_name(cell) for cell in header_row)


def _lookup_field_name(cell: object) -> str | None:
    if cell is None:
        return None

    return EWRB_HEADER_NORMALIZATION.get(str(cell).strip())


def _cell_to_string(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    return str(value)


def _load_workbook(body: bytes) -> Workbook:
    try:
        return openpyxl.load_workbook(BytesIO(body), read_only=True, data_only=True)
    except Exception as e:
        raise FetchError(
            "failed to open EWRB workbook",
            source_id=SOURCE_ID,
            dataset_id=EWRB_DATASET_ID,
            operation="open-workbook",
        ) from e


def _get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise FetchError(
            f"workbook is missing required sheet {sheet_name!r}",
            source_id=SOURCE_ID,
            dataset_id=EWRB_DATASET_ID,
            operation="open-workbook",
        )

    return workbook[sheet_name]
