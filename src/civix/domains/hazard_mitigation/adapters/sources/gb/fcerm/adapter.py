"""England FCERM scheme-allocation source adapter."""

from __future__ import annotations

import hashlib
import json
import zipfile
from collections.abc import AsyncIterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from typing import Final

import httpx
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from civix.core.identity.models.identifiers import DatasetId, Jurisdiction, SnapshotId, SourceId
from civix.core.mapping.parsers import slugify, str_or_none
from civix.core.ports.errors import FetchError
from civix.core.ports.models.adapter import FetchResult
from civix.core.snapshots.models.snapshot import RawRecord, SourceSnapshot
from civix.core.temporal import Clock, utc_now
from civix.domains.hazard_mitigation.adapters.sources.gb.fcerm.schema import (
    CEREMONIAL_COUNTY_FIELD,
    ENGLAND_FCERM_REQUIRED_FIELDS,
    INDICATIVE_GOVERNMENT_INVESTMENT_FIELD,
    LEAD_AUTHORITY_FIELD,
    PROJECT_NAME_FIELD,
)

SOURCE_ID: Final[SourceId] = SourceId("environment-agency")
GB_ENGLAND_JURISDICTION: Final[Jurisdiction] = Jurisdiction(country="GB", region="England")
ENGLAND_FCERM_SCHEMES_DATASET_ID: Final[DatasetId] = DatasetId("fcerm-schemes-2026-2027")
ENGLAND_FCERM_PROGRAMME_YEAR: Final[str] = "2026/27"
ENGLAND_FCERM_SCHEMES_SHEET_NAME: Final[str] = "List of schemes"
ENGLAND_FCERM_PUBLICATION_URL: Final[str] = (
    "https://www.gov.uk/government/publications/"
    "programme-of-flood-and-coastal-erosion-risk-management-schemes"
)
DEFAULT_RESOURCE_URL: Final[str] = (
    "https://assets.publishing.service.gov.uk/media/"
    "69bd49c913101e99087049ac/"
    "List_of_schemes_invested_in_between_April_2026_and_March_2027_V2.xlsx"
)
ENGLAND_FCERM_SOURCE_SCOPE: Final[str] = (
    "Environment Agency FCERM schemes receiving indicative government investment "
    "between April 2026 and March 2027. The default direct XLSX asset URL is pinned "
    "for reproducibility; if GOV.UK republishes the spreadsheet under a rotated asset "
    "URL, updating that URL is source-version maintenance."
)
ENGLAND_FCERM_SOURCE_RECORD_ID_POLICY: Final[str] = (
    "Derived from project name, lead risk management authority, ceremonial county, "
    "and workbook row number because the source workbook does not publish stable "
    "scheme identifiers. Mid-year row reordering can churn ids."
)


@dataclass(frozen=True, slots=True)
class _DecodedSchemeRow:
    row_number: int
    raw_data: Mapping[str, object]


@dataclass(frozen=True, slots=True)
class EnglandFcermFetchConfig:
    """Runtime fetch options for the year-pinned FCERM XLSX."""

    client: httpx.AsyncClient
    clock: Clock = field(default=utc_now)
    resource_url: str = DEFAULT_RESOURCE_URL


@dataclass(frozen=True, slots=True)
class EnglandFcermSchemesAdapter:
    """Fetches the England FCERM 2026/27 scheme-allocation workbook."""

    fetch_config: EnglandFcermFetchConfig

    @property
    def source_id(self) -> SourceId:
        return SOURCE_ID

    @property
    def dataset_id(self) -> DatasetId:
        return ENGLAND_FCERM_SCHEMES_DATASET_ID

    @property
    def jurisdiction(self) -> Jurisdiction:
        return GB_ENGLAND_JURISDICTION

    async def fetch(self) -> FetchResult:
        fetched_at = self.fetch_config.clock()
        content = await _fetch_bytes(self.fetch_config)
        content_hash = hashlib.sha256(content).hexdigest()
        rows = _parse_workbook(content)
        snapshot = _build_snapshot(
            fetched_at=fetched_at,
            record_count=len(rows),
            source_url=self.fetch_config.resource_url,
            content_hash=content_hash,
        )

        return FetchResult(
            snapshot=snapshot,
            records=_stream_records(snapshot=snapshot, rows=rows),
        )


async def _fetch_bytes(fetch: EnglandFcermFetchConfig) -> bytes:
    try:
        response = await fetch.client.get(fetch.resource_url)

        response.raise_for_status()
    except httpx.HTTPError as e:
        raise _fetch_error(
            f"failed to read FCERM XLSX from {fetch.resource_url}",
            operation="fetch-xlsx",
        ) from e

    return response.content


def _parse_workbook(content: bytes) -> tuple[_DecodedSchemeRow, ...]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, zipfile.BadZipFile) as e:
        raise _fetch_error("invalid FCERM XLSX response", operation="parse-xlsx") from e

    try:
        if ENGLAND_FCERM_SCHEMES_SHEET_NAME not in workbook.sheetnames:
            raise _fetch_error(
                f"missing {ENGLAND_FCERM_SCHEMES_SHEET_NAME!r} sheet",
                operation="parse-xlsx",
            )

        worksheet = workbook[ENGLAND_FCERM_SCHEMES_SHEET_NAME]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        headers = _validate_headers(header_row)
        rows: list[_DecodedSchemeRow] = []

        for row_number, row in enumerate(rows_iter, start=2):
            values = tuple(row)

            if _is_blank_row(values):
                continue

            rows.append(_decode_row(headers=headers, values=values, row_number=row_number))

        if not rows:
            raise _fetch_error("FCERM workbook contains no scheme rows", operation="parse-xlsx")

        return tuple(rows)
    finally:
        workbook.close()


def _validate_headers(header_row: Sequence[object] | None) -> tuple[str, ...]:
    if header_row is None:
        raise _fetch_error("missing FCERM workbook header", operation="parse-xlsx")

    headers = tuple(str_or_none(value) or "" for value in header_row)

    if headers != ENGLAND_FCERM_REQUIRED_FIELDS:
        raise _fetch_error(
            "FCERM workbook header does not match required fields",
            operation="parse-xlsx",
        )

    return headers


def _decode_row(
    *,
    headers: tuple[str, ...],
    values: tuple[object, ...],
    row_number: int,
) -> _DecodedSchemeRow:
    if len(values) != len(headers):
        raise _fetch_error(f"malformed FCERM workbook row {row_number}", operation="parse-xlsx")

    row = dict(zip(headers, values, strict=True))

    for field_name in ENGLAND_FCERM_REQUIRED_FIELDS:
        if str_or_none(row.get(field_name)) is None:
            raise _fetch_error(
                f"FCERM workbook row {row_number} missing required field {field_name!r}",
                operation="parse-xlsx",
            )

    amount_value = row[INDICATIVE_GOVERNMENT_INVESTMENT_FIELD]

    if isinstance(amount_value, bool) or not isinstance(amount_value, int | float | str):
        raise _fetch_error(
            f"FCERM workbook row {row_number} has invalid funding amount",
            operation="parse-xlsx",
        )

    return _DecodedSchemeRow(row_number=row_number, raw_data=row)


def _is_blank_row(values: tuple[object, ...]) -> bool:
    return all(str_or_none(value) is None for value in values)


async def _stream_records(
    *,
    snapshot: SourceSnapshot,
    rows: tuple[_DecodedSchemeRow, ...],
) -> AsyncIterable[RawRecord]:
    for row in rows:
        yield _build_record(snapshot=snapshot, row=row)


def _build_record(
    *,
    snapshot: SourceSnapshot,
    row: _DecodedSchemeRow,
) -> RawRecord:
    row_dict = dict(row.raw_data)
    source_record_id = _source_record_id(row_dict, row_number=row.row_number)
    record_hash = hashlib.sha256(json.dumps(row_dict, sort_keys=True).encode("utf-8")).hexdigest()

    try:
        return RawRecord(
            snapshot_id=snapshot.snapshot_id,
            raw_data=row_dict,
            source_record_id=source_record_id,
            source_updated_at=None,
            record_hash=record_hash,
        )
    except ValidationError as e:
        raise _fetch_error("raw FCERM record failed validation", operation="stream-records") from e


def _source_record_id(row: Mapping[str, object], *, row_number: int) -> str:
    maybe_parts = (
        _source_id_part(row.get(PROJECT_NAME_FIELD)),
        _source_id_part(row.get(LEAD_AUTHORITY_FIELD)),
        _source_id_part(row.get(CEREMONIAL_COUNTY_FIELD)),
        f"row-{row_number}",
    )

    if any(part is None for part in maybe_parts):
        raise _fetch_error(
            f"FCERM workbook row {row_number} cannot produce a source record id",
            operation="stream-records",
        )

    return ":".join(part for part in maybe_parts if part is not None)


def _source_id_part(value: object) -> str | None:
    text = str_or_none(value)

    if text is None:
        return None

    return slugify(text)


def _build_snapshot(
    *,
    fetched_at: datetime,
    record_count: int,
    source_url: str,
    content_hash: str,
) -> SourceSnapshot:
    snapshot_id = SnapshotId(
        f"{SOURCE_ID}:{ENGLAND_FCERM_SCHEMES_DATASET_ID}:{fetched_at.isoformat()}"
    )

    return SourceSnapshot(
        snapshot_id=snapshot_id,
        source_id=SOURCE_ID,
        dataset_id=ENGLAND_FCERM_SCHEMES_DATASET_ID,
        jurisdiction=GB_ENGLAND_JURISDICTION,
        fetched_at=fetched_at,
        record_count=record_count,
        source_url=source_url,
        fetch_params={
            "publication_url": ENGLAND_FCERM_PUBLICATION_URL,
            "programme_year": ENGLAND_FCERM_PROGRAMME_YEAR,
            "resource_url": source_url,
            "sheet_name": ENGLAND_FCERM_SCHEMES_SHEET_NAME,
            "source_record_id_policy": ENGLAND_FCERM_SOURCE_RECORD_ID_POLICY,
        },
        content_hash=content_hash,
    )


def _fetch_error(message: str, *, operation: str) -> FetchError:
    return FetchError(
        message,
        source_id=SOURCE_ID,
        dataset_id=ENGLAND_FCERM_SCHEMES_DATASET_ID,
        operation=operation,
    )
